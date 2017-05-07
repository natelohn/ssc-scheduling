import csv, os, random
from core import Stapher, Shift, Schedule, constants
from constraint import Problem, Variable, Domain
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Color, PatternFill, Font, Border, Side


def get_who_has_car(staph, file):
	car_file = open(file)
	for name in car_file:
		name = name[:-1]
		for stapher in staph['all-staph']:
			if stapher.name == name:
				stapher.has_car = True
	car_file.close()

def get_special_shift_rankings(staph, file):
	special_shift_file = open(file)
	csv_file = csv.reader(special_shift_file)
	for line in csv_file:
		name = line[0]
		special_shift_preferences = line[1:]
		for stapher in staph['all-staph']:
			if stapher.name == name:
				stapher.special_shift_preferences = special_shift_preferences
	special_shift_file.close()

def get_off_day_pairs(staph,file):
	off_day_file = open(file)
	csv_file = csv.reader(off_day_file)
	for line in csv_file:
		name = line[0]
		off_day_pairs = line[1:]
		for stapher in staph['all-staph']:
			if stapher.name == name:
				stapher.pairs_to_avoid = off_day_pairs
	off_day_file.close()

def get_staph_from_csv_file(file):
	staph_file = open(file)
	csv_staph_file = csv.reader(staph_file)
	staph = {}
	staph['all-staph'] = []
	for staph_info in csv_staph_file:
		name = staph_info[0]
		summers = staph_info[1]
		class_year = staph_info[2]
		gender = staph_info[3]
		positions = staph_info[4:]
		stapher = Stapher(name,summers,class_year,gender,positions)
		staph['all-staph'].append(stapher)
		for position in positions:
			if position in staph:
				staph[position].append(stapher)
			else:
				staph[position] = [stapher]
	staph_file.close()
	get_who_has_car(staph,'../input/2017/info/has-car.csv')
	get_special_shift_rankings(staph,'../input/2017/info/special-shift-rankings.csv')
	get_off_day_pairs(staph,'../input/2017/info/day-off-pairs.csv')
	return staph

def get_shifts_from_csv_file(filename, is_programming):
	shifts_in_file = []
	shifts_file = open(filename)
	csv_shifts_file = csv.reader(shifts_file)
	for shift_info in csv_shifts_file:
		day = int(shift_info[0])
		title = shift_info[1]
		start = float(shift_info[2])
		end = float(shift_info[3])
		ammount = int(shift_info[4])
		for i in range(0,ammount):
			shift = Shift(day, title, start, end, is_programming)
			shifts_in_file.append(shift)
	shifts_file.close()
	return shifts_in_file

def get_shifts(directory):
	shifts = {}
	shifts['all-shifts'] = []
	for category in os.listdir(directory):
		category_dir = directory + '/' + category
		if 'DS' not in category_dir: # need to remove this after putting git magic in
			shifts[category] = {}
			shifts['all-' + category + '-shifts'] = []
			for worker_group in os.listdir(category_dir):
				worker_dir =  category_dir + '/' + worker_group
				if 'DS' not in worker_dir:
					types_for_worker_group = []
					shifts[category][worker_group] = []
					shifts['all-' + worker_group + '-shifts'] = []
					for shift_file in os.listdir(worker_dir):
						if 'DS' not in shift_file:
							file = worker_dir + '/' + shift_file
							shift_type = shift_file[5:-4]
							shifts[shift_type] = []
							if category == 'programming':
								shifts_of_type = get_shifts_from_csv_file(file, True)
							else:
								shifts_of_type = get_shifts_from_csv_file(file, False)
							shifts[category][worker_group] += [shift_type]
							shifts['all-' + category + '-shifts'] += shifts_of_type
							shifts['all-' + worker_group + '-shifts'] += shifts_of_type
							shifts[shift_type] += shifts_of_type
							shifts['all-shifts'] += shifts_of_type
							types_for_worker_group.append(shift_type)
					if worker_group in shifts:
						shifts[worker_group] += types_for_worker_group
					else:
						shifts[worker_group] = types_for_worker_group
	return shifts

def get_time_from_txt(time_txt):
	split = time_txt.index(':')
	hour = int(time_txt[:split])
	minute_txt = time_txt[split:-2]
	is_am = time_txt[-2:] == 'am'
	if minute_txt == ':00':
		minute = 0
	elif minute_txt == ':15':
		minute = 0.25
	elif minute_txt == ':20':
		minute = 0.33
	elif minute_txt == ':30':
		minute = 0.5
	elif minute_txt == ':40':
		minute = 0.66
	elif minute_txt == ':45':
		minute = 0.75
	if not is_am and hour != 12: 
		hour += 12
	return hour + minute

def txt_matches_shift(shift,shift_txt,day):
	shift_txt_is_programming = shift_txt[len(shift_txt) - 1:] == 'p'
	if shift_txt_is_programming:
		shift_txt = shift_txt[:-1]
	time_index = shift_txt.index('(')
	time_txt = shift_txt[time_index:]
	start_index = time_txt.index('-')
	shift_title = shift_txt[:time_index]
	start_time_txt = time_txt[1:start_index]
	end_time_txt = time_txt[start_index + 1:-1]
	start_time = get_time_from_txt(start_time_txt)
	end_time = get_time_from_txt(end_time_txt)
	if shift.title != shift_title:
		# if 'Grimm' in shift.title and 'Grimm' in shift_txt:
			# print shift.title, 'title not', shift_title
			# print len(shift.title), len(shift_title)
		return False
	if shift.day != day:
		# if 'Grimm' in shift.title and 'Grimm' in shift_txt:
			# print shift.day, 'day not', day
		return False
	if shift.start != start_time:
		# if 'Grimm' in shift.title and 'Grimm' in shift_txt:
			# print shift.start_time, 'start not', start_time
		return False
	if shift.end != end_time:
		# if 'Grimm' in shift.title and 'Grimm' in shift_txt:
			# print shift.end_time, 'end not', end_time
		return False
	if shift.is_programming != shift_txt_is_programming:
		# if 'Grimm' in shift.title and 'Grimm' in shift_txt:
		# 	print shift.is_programming, 'programming not', shift_txt_is_programming
		return False
	# if 'Grimm' in shift.title and 'Grimm' in shift_txt:
		# print shift_txt,'->',shift
	return True

def get_free_shift_from_txt(day,shift_txt,shifts):
	for shift in shifts['all-shifts']:
		if not shift.covered:
			if txt_matches_shift(shift,shift_txt,day):
				return shift
	print 'No free shift that matches',shift_txt

def get_schedules_from_txt_files(staph,shifts):
	days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday','Saturday']
	for stapher in staph['all-staph']:
		schedule_txt_file = open('../input/2017/schedules/txt_files/' + stapher.name + '.txt')
		csv_schedule_txt_file = csv.reader(schedule_txt_file)
		for line in csv_schedule_txt_file:
			day_txt = line[0]
			days_shifts = line[1:]
			day = days.index(day_txt)
			for shift_txt in days_shifts:
				shift = get_free_shift_from_txt(day,shift_txt,shifts)
				if stapher.free_during_shift(shift):
					stapher.add_shift(shift)
		schedule_txt_file.close()


# This is a method to see which days position groups can not take off. (i.e. Ski Dock can't be TueWay)
def get_off_day_restrictions(staph, shifts):
	restricted_off_days = {}
	for position in shifts['programming']:
		if position in staph:
			programming = 'all-' + position + '-shifts'
			ammount_of_shifts_at_a_given_time = {}
			for shift in shifts[programming]:
				time_key = str(shift.day) + ' ' + str(shift.start) + '-' +  str(shift.end) # Does not account for overlapping times
				if time_key in ammount_of_shifts_at_a_given_time:
					ammount_of_shifts_at_a_given_time[time_key] += 1
				else:
					ammount_of_shifts_at_a_given_time[time_key] = 1
				if len(staph[position]) - ammount_of_shifts_at_a_given_time[time_key] == 0:
					if position in restricted_off_days:
						restricted_off_days[position].add(shift.day)
					else:
						restricted_off_days[position] = set([shift.day])
					if shift.start >= 16:
						restricted_off_days[position].add(shift.day + 1)
	# off_day_names = ['SatSu','SuMo','MoTue','TueWay','WeThur','StirFry','FriSat','SatSu']
	# for position in restricted_off_days:
	# 	print position, 'can not be...'
	# 	for day in restricted_off_days[position]:
	# 		print '	', off_day_names[day]
	return restricted_off_days

def passes_off_day_restrictions(stapher, off_day, off_day_restrictions):
	for position in stapher.positions:
		if position in off_day_restrictions:
			if off_day in off_day_restrictions[position]:
				return False
	if not stapher.free_during_time(off_day,0,24) or not stapher.free_during_time(off_day - 1,16,24):
		return False
	return True

def get_avg_cars(off_day_groups):
	total_cars = 0
	for group in off_day_groups:
		for stapher in off_day_groups[group]:
			if stapher.has_car:
				total_cars += 1
	avg_cars = total_cars / len(off_day_groups)
	return avg_cars

def off_day_group_score(group,avg_cars):
	count_of_positions = {}
	count_of_gender = {}
	count_of_class_years = {}
	position_overlap = 0
	returner_count = 0
	bad_pair_count = 0
	bad_pairs = []
	cars_in_group = 0
	for stapher in group:
		if stapher.has_car:
			cars_in_group += 1
		for other_stapher in group:
			if other_stapher.name in stapher.pairs_to_avoid:
				string_one = stapher.name + ',' + other_stapher.name
				string_two = other_stapher.name + ',' + stapher.name
				if string_one not in bad_pairs and string_two not in bad_pairs:
					bad_pair_count += 1
					bad_pairs.append(stapher.name + ',' + other_stapher.name)
		if stapher.is_returner():
			returner_count += 1
		if stapher.class_year in count_of_class_years:
			count_of_class_years[stapher.class_year] += 1
		else:
			count_of_class_years[stapher.class_year] = 1
		if stapher.gender in count_of_gender:
			count_of_gender[stapher.gender] += 1
		else:
			count_of_gender[stapher.gender] = 1
		for position in stapher.positions:
			if position != 'boat-dock-coordinator':
				if position in count_of_positions:
					count_of_positions[position] += 1
					position_overlap += count_of_positions[position]
				else:
					count_of_positions[position] = 1
	majority_class_year = 0
	for class_year in count_of_class_years:
		if count_of_class_years[class_year] >= majority_class_year:
			majority_class_year = count_of_class_years[class_year]
	if (len(group) - returner_count) < 0:
		returner_new_balence_score = returner_count - (len(group) - returner_count)
	else:
		returner_new_balence_score = (len(group) - returner_count) - returner_count
	if 'male' not in count_of_gender or 'female' not in count_of_gender:
		print 'NOT THIS ONE!'
		return 10000000
	if count_of_gender['female'] > count_of_gender['male']:
		majority_gender = 'female'
	else:
		majority_gender = 'male'
	gender_balence = count_of_gender[majority_gender] - (len(group) - count_of_gender[majority_gender])
	position_weight = position_overlap * 2
	bad_pair_weight = bad_pair_count
	car_score = abs(avg_cars - cars_in_group)
	gender_weight = 3
	returner_new_weigt = 5
	class_year_weight = 1
	car_weight = 1
	group_score = (car_score * car_weight) + (position_overlap * position_weight) + (bad_pair_count * bad_pair_weight) + (gender_balence * gender_weight) + (returner_new_balence_score * returner_new_weigt) + (majority_class_year * class_year_weight)
	return group_score

def get_has_been_stirfry(stirfry):
	filename = '../input/2017/info/has-been-stirfry.csv'
	prev_stirfry_file = open(filename)
	csv_file = csv.reader(prev_stirfry_file)
	prev_stirfry_names = []
	for line in csv_file:
		name = line[0]
		prev_stirfry_names.append(name)
	stirfry_repeats = []
	for stapher in stirfry:
		if stapher in prev_stirfry_names:
			stirfry_repeats.append(stapher)
	prev_stirfry_file.close()
	return stirfry_repeats


def get_off_days_score(off_day_groups):
	scores = {}
	total_score = 0
	avg_cars = get_avg_cars(off_day_groups)
	for group in off_day_groups:
		total_score += off_day_group_score(off_day_groups[group],avg_cars)
	stirfry_names = []
	for stapher in off_day_groups[5]:
		stirfry_names.append(stapher.name)
	stirfry_repeats = get_has_been_stirfry(stirfry_names)
	stirfry_repeats_score = len(stirfry_repeats) * len(stirfry_repeats)
	total_score += stirfry_repeats_score
	return total_score

def make_off_day_csv(off_day_group):
	directory = '../output/2017/masters'
	if not os.path.exists(directory):
		os.makedirs(directory)
	filename = directory + '/off_days.csv'
	new_file = open(filename, 'w')
	off_day_names = ['SatSu','SuMo','MoTue','TueWay','WeThur','StirFry','FriSat']
	txt = ''
	for day in off_day_group:
		txt += off_day_names[day]
		for stapher in off_day_group[day]:
			txt += ',' + stapher.name
		txt += '\n'
	new_file.write(txt)
	new_file.close()			

def get_row_from_position(position):
	if position == 'munchkins':
		return 2
	elif position == 'snoopers':
		return 3
	elif position == 'menehunes':
		return 4
	elif position == 'yahoos':
		return 5
	elif position == 'midorees':
		return 6
	elif position == 'suaves':
		return 7
	elif position == 'teens':
		return 8
	elif position == 'ski-dock':
		return 9
	elif position == 'chicken':
		return 10
	elif position in ['arts','crafts', 'music','theater']:
		return 11
	elif position in ['hiking','naturalist', 'kids-naturalist']:
		return 12
	elif position in ['sailing','tennis','climbing','yoga','volleyball']:
		return 13
	elif position in ['office','photo']:
		return 14
	elif position in ['staph-d','iic', 'kgc']:
		return 15
	print position, 'didnt return anything...'

def reset_off_day_xl():
	file = '../output/2017/masters/off-days-master.xlsx'
	wb = load_workbook(file)
	ws = wb.get_sheet_by_name('Off Days')
	for day in range(1,6):
		col = day + 1
		for row in range(2,16):
			cell = ws.cell(row=row,column=col)
			cell.value = ''
	wb.save(file)


def populate_off_day_xl(off_day_group):
	file = '../output/2017/masters/off-days-master.xlsx'
	wb = load_workbook(file)
	ws = wb.get_sheet_by_name('Off Days')
	for day in off_day_group:
		col = day + 1
		for stapher in off_day_group[day]:
			row = get_row_from_position(stapher.positions[0])
			cell = ws.cell(row=row,column=col)
			if cell.value == None:
				cell.value = stapher.name
			else:
				cell.value = cell.value + '\n' + stapher.name
	wb.save(file)


def make_off_day_xl(staph):
	reset_off_day_xl()
	off_day_group = get_groups_from_csv(staph)
	populate_off_day_xl(off_day_group)


def generate_random_off_day_groups(staph, off_day_restrictions):
	staph_to_be_chosen = list(staph['all-staph'])
	off_day_groups = {}
	tries = 0
	possible_off_days = [1,2,3,4,5]
	while len(staph_to_be_chosen) > 0:
		stapher = staph_to_be_chosen[random.randint(0, len(staph_to_be_chosen) - 1)]
		shortest_off_day = possible_off_days[0]
		if tries > 1000:
			staph_to_be_chosen = list(staph['all-staph'])
			off_day_groups = {}
			tries = 0
		if passes_off_day_restrictions(stapher, shortest_off_day, off_day_restrictions):
			if shortest_off_day in off_day_groups:
				off_day_groups[shortest_off_day].append(stapher)
			else:
				off_day_groups[shortest_off_day] = [stapher]
			staph_to_be_chosen.remove(stapher)
			possible_off_days.pop(0) #Removes the shortest off day from the front
			possible_off_days.append(shortest_off_day)#Place the shortest off day at the back
		tries += 1
	return off_day_groups

def schedule_off_days(best_group):
	for day in best_group:
		for stapher in best_group[day]:
			half_off_day = Shift(day - 1,'Off Day',16,23.5,False)
			off_day = Shift(day,'Off Day',6,23.5,False)
			stapher.add_shift(half_off_day)
			stapher.add_shift(off_day)

def make_off_day_statistics(staph,shifts):
	best_group = get_groups_from_csv(staph)
	off_day_restrictions = get_off_day_restrictions(staph, shifts)
	score = get_off_days_score(best_group)
	statistics = 'SCORE: ' + str(score)
	errors = ''
	off_day_names = ['SatSu','SuMo','MoTue','TueWay','WeThur','StirFry','FriSat']
	avg_cars = get_avg_cars(best_group)
	for day in best_group:
		genders = {}
		summers = {}
		positions = {}
		class_years = {}
		bad_pairs = []
		has_car_on_off_day = []
		group = best_group[day]
		if day == 5:
			group_names = []
			for stapher in group:
				group_names.append(stapher.name)
			stirfry_repeats = get_has_been_stirfry(group_names)
			stirfry_repeats_txt = ''
			if len(stirfry_repeats) > 0:
				stirfry_repeats_txt = str(len(stirfry_repeats)) + ' staphers have/has been stirfry before:\n 	'
				for name in stirfry_repeats:
					stirfry_repeats_txt += name + ','
				stirfry_repeats_txt = stirfry_repeats_txt[:-1] + '\n'
		for stapher in group:
			for position in stapher.positions:
				if position in off_day_restrictions:
					if day in off_day_restrictions[position]:
						errors  += '\n' + stapher.name + ' CAN NOT BE ' + off_day_names[day] + ' BECAUSE THEY ARE ' + position + '\n'
			if stapher.has_car:
				has_car_on_off_day.append(stapher.name)
			if stapher.gender in genders:
				genders[stapher.gender].append(stapher)
			else:
				genders[stapher.gender] = [stapher]
			if stapher.summers in summers:
				summers[stapher.summers].append(stapher)
			else:
				summers[stapher.summers] = [stapher]
			if stapher.class_year in class_years:
				class_years[stapher.class_year].append(stapher)
			else:
				class_years[stapher.class_year] = [stapher]
			for position in stapher.positions:
				if position in positions:
					positions[position].append(stapher)
				else:
					positions[position] = [stapher]
			for other_stapher in best_group[day]:
				if other_stapher.name in stapher.pairs_to_avoid:
					string_one = stapher.name + ',' + other_stapher.name
					string_two = other_stapher.name + ',' + stapher.name
					if string_one not in bad_pairs and string_two not in bad_pairs:
						bad_pairs.append(stapher.name + ',' + other_stapher.name)
		statistics += '\n=========== ' + off_day_names[day] + ' ===========\n'
		stapher_count = len(group)
		statistics += str(stapher_count) + ' staphers\n'
		group_score = off_day_group_score(group,avg_cars)
		if day == 5:
			statistics += stirfry_repeats_txt
		statistics += 'GROUP SCORE:' + str(group_score) + '\n'
		statistics += '	Gender Balence: ' 
		for gender in genders:
			statistics += str(len(genders[gender])) + ' ' +  gender + ', '
		returners = 0
		returner_males = 0
		new_males = 0
		for summer in summers:
			if summer > 0:
				returners += len(summers[summer])
			for stapher in summers[summer]:
				if stapher.gender == 'male':
					if summer > 0:
						returner_males += 1
					else:
						new_males += 1
		new = stapher_count - returners
		returner_females = returners - returner_males
		new_females = new - new_males
		statistics += '\n 	Returner/New Balence: ' + str(new) + ' new & ' + str(returners) + ' returning.'
		statistics += '\n 		' + str(new_females) + ' new females & ' + str(returner_females) + ' returning females.'
		statistics += '\n 		' + str(new_males) + ' new males & ' + str(returner_males) + ' returning males.'
		statistics += '\n 	Class Year:\n'
		for class_year in class_years:
			males_in_class = 0
			for stapher in class_years[class_year]:
				if stapher.gender == 'male':
					males_in_class += 1
			females_in_class = len(class_years[class_year]) - males_in_class
			statistics += '		' + str(len(class_years[class_year])) + ' in ' + str(class_year) + ': ' + str(males_in_class) + ' males, ' + str(females_in_class) + ' females.\n'
		statistics += '\n 	' + str(len(has_car_on_off_day)) + ' people have cars: '
		for name in has_car_on_off_day:
			statistics += name + ','
		statistics += '\n 	' + str(len(bad_pairs)) + ' bad pair(s): '
		for pair in bad_pairs:
			statistics += pair + ', '
		statistics += '\n 	Positions:'
		for position in positions:
			statistics += '\n 		' + str(len(positions[position])) + ' ' + position
	statistics = errors + statistics
	print '		OFF DAY STATISTICS MADE'
	directory = '../output/2017/info'
	if not os.path.exists(directory):
		os.makedirs(directory)
	filename = directory + '/off_day_stats.txt'
	new_file = open(filename, 'w')
	new_file.write(statistics)
	new_file.close()

def improve_off_day_groups(off_day_groups,group1,group2,score_to_beat,off_day_restrictions):
	for stapher in off_day_groups[group1]:
		if passes_off_day_restrictions(stapher, group2, off_day_restrictions):
			off_day_groups[group1].remove(stapher)
			off_day_groups[group2].append(stapher)
			current_score = get_off_days_score(off_day_groups)
			if current_score < score_to_beat:
				return off_day_groups
			else:
				off_day_groups[group2].remove(stapher)
				off_day_groups[group1].append(stapher)
	return off_day_groups

def get_better_off_day_groups(off_day_groups,staph,off_day_restrictions):
	current_score = get_off_days_score(off_day_groups)
	group1 = random.randint(1,len(off_day_groups))
	group2 = group1
	while group2 == group1:
		group2 = random.randint(1,len(off_day_groups))
	if len(off_day_groups[group1]) > len(off_day_groups[group2]):
		off_day_groups = improve_off_day_groups(off_day_groups,group1,group2,current_score,off_day_restrictions)
	else:
		off_day_groups = improve_off_day_groups(off_day_groups,group2,group1,current_score,off_day_restrictions)
	return off_day_groups

def generate_best_groups(staph, off_day_restrictions):
	best_scoring_groups = []
	off_day_groups = generate_random_off_day_groups(staph, off_day_restrictions)
	no_improvment = 0
	reset_at = 50
	stop_at =  5000
	target_score = 161
	new_score = target_score + 1
	reset_count = 0
	lowest_scores = []
	reset_count = 0
	lowest_score = 1000
	best_off_day_group = []
	while stop_at > reset_count:
		old_score = get_off_days_score(off_day_groups)
		off_day_groups = get_better_off_day_groups(off_day_groups,staph,off_day_restrictions)
		new_score = get_off_days_score(off_day_groups)
		if old_score == new_score:
			no_improvment += 1
		else:
			no_improvment = 0
			print 'Attempt:',reset_count,'- Current Score:',new_score,'- Best Score:',lowest_score
			if new_score < lowest_score:
				lowest_score = new_score
				best_off_day_group = off_day_groups
		if no_improvment == reset_at:
			print '------------------'
			off_day_groups = generate_random_off_day_groups(staph, off_day_restrictions)
			no_improvment = 0
			reset_count += 1
			lowest_scores.append(new_score)
	return best_off_day_group

def get_groups_from_csv(staph):
	off_days = open('../input/2017/schedules/off_days.csv')
	csv_off_day_file = csv.reader(off_days)
	off_day_names = ['SatSu','SuMo','MoTue','TueWay','WeThur','StirFry','FriSat']
	off_day_groups = {}
	for line in csv_off_day_file:
		off_day = off_day_names.index(line[0])
		names = line[1:]
		for stapher in staph['all-staph']:
			if stapher.name in names:
				if off_day in off_day_groups:
					off_day_groups[off_day].append(stapher)
				else:
					off_day_groups[off_day] = [stapher]
	off_days.close()
	return off_day_groups

def generate_off_days(staph,shifts):
	off_day_restrictions = get_off_day_restrictions(staph, shifts)
	off_day_groups = generate_best_groups(staph, off_day_restrictions)
	make_off_day_csv(off_day_groups)

def upload_off_days(staph,shifts):
	off_day_groups = get_groups_from_csv(staph)
	schedule_off_days(off_day_groups)
	

def count_of_breakfast_chefs(stapher):
	chef_count = 0
	for shift in stapher.all_programming_shifts():
		if shift.title == 'Breakfast Chef':
			chef_count += 1
	return chef_count

def count_of_chichen_splits(stapher):
	split_count = 0
	for day in stapher.schedule.all_shifts:
		count_of_chicken_shifts = 0
		for shift in stapher.schedule.get_day_schedule(day):
			if shift.title == 'Chicken' or shift.title == 'Breakfast Chef':
				count_of_chicken_shifts += 1
		if count_of_chicken_shifts > 1:
			split_count += 1
	return split_count

def get_staphers_programming_score(stapher,staph_group):
	avg_prgramming = get_avg_programming_for_group(staph_group)
	score = abs(avg_prgramming - stapher.programming_hours())
	score += stapher.meal_break_violations()
	if 'ski-dock' in stapher.positions:
		for length in stapher.get_lengths_without_break():
			if length > 3:
				score += 1
	if 'chicken' in stapher.positions:
		chef_counts = count_of_breakfast_chefs(stapher)
		score += abs(1 - chef_counts)
		chichen_splits = count_of_chichen_splits(stapher)
		score += abs(1 - chichen_splits)
	return score

def get_avg_programming_for_group(staph_group):
	total_programming_hours = 0
	for stapher in staph_group:
		total_programming_hours += stapher.programming_hours()
	return total_programming_hours / len(staph_group)

def get_programming_score(staph_group,programming_shifts):
	score = 0
	total_programming_hours = 0
	kids_groups = ['munchkins','snoopers','menehunes','yahoos','midorees','suaves','teens']
	is_kids_group = False
	for stapher in staph_group:
		if stapher.positions[0] in kids_groups:
			is_kids_group = True
		score += get_staphers_programming_score(stapher,staph_group)
	if is_kids_group:
		for shift in programming_shifts:
			all_shifts = get_total_overlapping_shifts(shift,programming_shifts)
			all_shifts.append(shift)
			male_on_shift = False
			female_on_shift = False
			for overlapping_shift in all_shifts:
				# print shift,overlapping_shift
				if overlapping_shift.stapher.gender == 'male':
					male_on_shift = True
				elif overlapping_shift.stapher.gender == 'female':
					female_on_shift = True
			if not male_on_shift or not female_on_shift:
				score += 1
	return score

def get_bests_and_worst_scores(staph_group):
	best_score = 1000
	worst_score = 0
	best_schedule = None
	worst_schedule = None
	for stapher in staph_group:
		stapher_score = get_staphers_programming_score(stapher,staph_group)
		if stapher_score < best_score:
			best_score = stapher_score
			best_schedule = stapher
		if stapher_score > worst_schedule:
			worst_score = stapher_score
			worst_schedule = stapher
	return [best_schedule,worst_schedule]


def find_improved_swaps(staph_group,giving_stapher,recieving_stapher,programming_shifts,score_to_beat):
	for shift in giving_stapher.all_shifts():
		if shift in programming_shifts:
			if recieving_stapher.free_during_shift(shift):
				giving_stapher.remove_shift(shift)
				recieving_stapher.add_shift(shift)
				new_score = get_programming_score(staph_group,programming_shifts)
				if new_score < score_to_beat:
					score_to_beat = new_score
				else:
					recieving_stapher.remove_shift(shift)
					giving_stapher.add_shift(shift)
				# print get_programming_score(staph_group,programming_shifts)
					

def get_staphers_free_during_shift(staphers,shift):
	free_staphers = []
	for stapher in staphers:
		if stapher.free_during_shift(shift):
			free_staphers.append(stapher)
	return free_staphers

def make_programming_stats(staph):
	stats = ''
	day_names = ['Sun','Mon','Tues','Wed','Thur','Fri','Sat']
	kids_groups = ['munchkins','snoopers','menehunes','yahoos','midorees','suaves','teens']
	chicken_to_print = 0
	ski_dock_to_print = 0
	for position in staph: 
		if position in shifts and position not in 'all-staph':
			staphers = staph[position]
			all_programming_shifts = shifts['all-' + position + '-shifts']
			programming_hours = 0
			uncoved_shifts = []
			shifts_without_both_genders = []
			for shift in all_programming_shifts:
				programming_hours += shift.length
				if not shift.covered:
					uncoved_shifts.append(shift)
				elif position in kids_groups:
					all_shifts = get_total_overlapping_shifts(shift,all_programming_shifts)
					male_on_shift = shift.stapher.gender == 'male'
					female_on_shift = shift.stapher.gender == 'female'
					for other_shift in all_shifts:
						if other_shift.stapher.gender == 'male':
							male_on_shift = True
						if other_shift.stapher.gender == 'female':
							female_on_shift = True
					if not male_on_shift or not female_on_shift:
						if all_shifts not in shifts_without_both_genders:
							shifts_without_both_genders.append(all_shifts)
			stats += '\n=============================' + position.upper() +'===============================\n'	
			stats += str(len(all_programming_shifts)) + ' TOTAL SHIFTS, ' + str(len(uncoved_shifts)) + ' SHIFT(S) UNCOVERED!\n'
			if position in kids_groups:
				stats = stats[:-2] + ', ' + str(len(shifts_without_both_genders)) + ' SHIFTS WITHOUT BOTH GENDERS!\n'
			place_holder = len(stats)
			if len(uncoved_shifts) > 0:
				stats += 'UNCOVERED SHIFT(S):\n'
				for shift in uncoved_shifts:
					stats +=  '	' + str(shift) + '\n'	
			if len(shifts_without_both_genders) > 0:
				stats += 'SHIFT(S) WITHOUT BOTH GENDERS:\n'
				for shift in shifts_without_both_genders:
					stats +=  '	' + str(shift[0]) + '\n'
			total_three_hour_plus_ski_dock_shifts = 0
			total_chicken_splits = 0
			for stapher in staphers:
				stats += '	' + stapher.name + ', ' + str(len(stapher.all_programming_shifts())) + ' programming shifts, ' +  str(stapher.programming_hours()) + ' programming hours\n'
				if 'chicken' in stapher.positions:
					chef_counts = count_of_breakfast_chefs(stapher)
					chichen_splits = count_of_chichen_splits(stapher)
					stats = stats[:-1] + ', '+ str(chef_counts) + ' Breakfast Chef(s), ' + str(chichen_splits) + ' Chicken Split(s)\n'
					total_chicken_splits += chichen_splits
				three_hour_plus_shifts = 0
				for length in stapher.get_lengths_without_break():
					if length > 3:
						three_hour_plus_shifts += 1
						total_three_hour_plus_ski_dock_shifts += 1
				if three_hour_plus_shifts > 0 and 'ski-dock' in position:
					stats = stats[:-1] + ', '+ str(three_hour_plus_shifts) + ' SHIFT(S) LONGER THAN 3 HOURS!\n'
				for day in range(0,7):
					stats += '		' + day_names[day] + ': ' + str(len(stapher.get_programming_shifts_by_day(day))) + ' shifts, ' +  str(stapher.get_programming_hours_by_day(day))+ ' hours\n'
			if 'ski-dock' in position:
				stats = stats[:place_holder - 1] + ', ' + str(total_three_hour_plus_ski_dock_shifts) + ' SHIFTS LONGER THAN 3 HOURS!\n' + stats[place_holder:]
				ski_dock_to_print += total_three_hour_plus_ski_dock_shifts
			if 'chicken' in position:
				stats = stats[:place_holder - 1] + ', ' + str(total_chicken_splits) + ' CHCIKEN SPLIT(S)!\n' + stats[place_holder:]
				chicken_to_print += total_chicken_splits
	# print stats
	filename = '../output/2017/info/programming_stats.txt'
	new_file = open(filename, 'w')
	new_file.write(stats)
	new_file.close()
	return[chicken_to_print,ski_dock_to_print]

def automate_programming(staphers, shifts_to_fill):
	# This method is best is programming is placed first
	break_period = 0.5
	for shift in shifts_to_fill:
		free_staphers = get_staphers_free_during_shift(staphers,shift)
		if len(free_staphers) == 0:
			print 'NO STAPHERS CAN COVER',shift
			quit()
		chosen_stapher = free_staphers[random.randint(0,len(free_staphers) - 1)]
		for stapher in free_staphers:
			chosen_day_hours = chosen_stapher.get_programming_hours_by_day(shift.day)
			staphers_day_hours = stapher.get_programming_hours_by_day(shift.day)
			if stapher.programming_hours() < chosen_stapher.programming_hours():
				if staphers_day_hours < chosen_day_hours:
					chosen_stapher = stapher
				elif stapher.free_during_time(shift.day,shift.start - break_period, shift.start + break_period):
						chosen_stapher = stapher
			elif staphers_day_hours < chosen_day_hours:
				if stapher.free_during_time(shift.day,shift.start - break_period, shift.start + break_period):
					chosen_stapher = stapher
		chosen_stapher.add_shift(shift)

def programming(staph,shifts):
	positions = staph.keys()
	number_of_staphers = 1
	ordered_staph_groups = []
	while len(ordered_staph_groups) < len(positions):
		for position in positions:
			if len(staph[position]) == number_of_staphers:
				ordered_staph_groups.append(position)
		number_of_staphers += 1
	ordered_staph_groups.remove('all-staph')
	for position in ordered_staph_groups:
		staphers = staph[position]
		for shift_type in shifts[position]:
			shifts_to_fill = list(shifts[shift_type])
			# Remove shifts that have been placed already
			for shift in shifts_to_fill:
				if shift.covered:
					shifts_to_fill.remove(shift)
			automate_programming(staphers, list(shifts_to_fill))


						
	

def stapher_has_no_shift_of_type(stapher,preference,shifts):
	for shift in stapher.all_shifts():
		if shift in shifts[preference + '-shifts']:
			return False
		if 'hobart' in preference:
			if shift in shifts['breakfast-hobart-shifts'] or shift in shifts['lunch-hobart-shifts'] or shift in shifts['dinner-hobart-shifts']:
				return False
	return True
		
def get_top_ranked_special_shift(shifts,stapher):
	for preference in stapher.special_shift_preferences:
		if stapher_has_no_shift_of_type(stapher,preference,shifts):
			for ranked_shift in shifts[preference + '-shifts']:
				if stapher.free_during_shift(ranked_shift) and not ranked_shift.covered:
					return ranked_shift

def by_class_year_and_summers(s1, s2):
	score1 = s1.summers + (2020 - s1.class_year)
	score2 = s2.summers + (2020 - s2.class_year)
	return score2 - score1

def get_preference_of_shift(shift,all_preferences,shifts):
	for preference in all_preferences:
		if shift in shifts[preference]:
			return preference[:-7]
	
def get_counting_string(num):
	if num == 1:
		prefix = 'st'
	elif num == 2:
		prefix = 'nd'
	elif num == 3:
		prefix = 'rd'
	else:
		prefix = 'th'
	return str(num) + prefix

def make_special_shift_statistics(staph,shifts):
	stat_text = ''
	warnings = ''
	by_ranks = {}
	by_names = {}
	ranked_shifts_scheduled = []
	all_special_shifts = shifts['all-non-flexible-shifts']
	for stapher in staph['all-staph']:
		stapher_ranks = {}
		for rank,preference in enumerate(stapher.special_shift_preferences):
			for shift in shifts[preference + '-shifts']:
				if shift in stapher.all_shifts():
					ranked_shifts_scheduled.append(shift)
					if rank in stapher_ranks:
						stapher_ranks[rank].append(shift)
					else:
						stapher_ranks[rank] = [shift]
					if rank in by_ranks:
						by_ranks[rank].append(shift)
					else:
						by_ranks[rank] = [shift]
		by_names[stapher.name] = stapher_ranks
	ranked_and_not_scheduled = []
	who_ranked_shifts = []
	unranked = []
	for shift in all_special_shifts:
		if shift not in ranked_shifts_scheduled:
			shift_ranked = False
			for stapher in staph['all-staph']:
				for preference in stapher.special_shift_preferences:
					if shift in shifts[preference + '-shifts']:
						shift_ranked = True
						if shift not in ranked_and_not_scheduled:
							ranked_and_not_scheduled.append(shift)
						if stapher_has_no_shift_of_type(stapher,preference,shifts):
							string = stapher.name + ' ranked ' + str(shift) + ' ' + get_counting_string(stapher.special_shift_preferences.index(preference) + 1)
							who_ranked_shifts.append(string)
			if not shift_ranked:
				unranked.append(shift)
	p = float(len(unranked)) / float(len(all_special_shifts)) * 100
	stat_text += str(len(unranked)) + '/' + str(len(all_special_shifts)) + ' (' + str(p) + '%) placed were not ranked by any staphers. *not accounting for flexible shifts.\n'
	p = float(len(ranked_shifts_scheduled)) / float(len(all_special_shifts)) * 100
	stat_text += str(len(ranked_shifts_scheduled)) + '/'+ str(len(all_special_shifts))+' ('+ str(p)+'%) placed were given to staphers that rankeed them. *not accounting for flexible shifts.\n'
	p = float(len(ranked_and_not_scheduled)) / float(len(all_special_shifts)) * 100
	stat_text += str(len(ranked_and_not_scheduled))+ '/'+ str(len(all_special_shifts)) + ' ('+ str(p) + '%) placed were ranked and given to staphers that did not rank them. *not accounting for flexible shifts.\n'
	for text in who_ranked_shifts:
		stat_text += '	' + text + '\n'
	stat_text += '-------------------------------------------------------------\n'
	scheduled_in_range = 0
	for rank in by_ranks:
		scheduled_in_range += len(by_ranks[rank])
		p = float(scheduled_in_range) / float(len(all_special_shifts)) * 100
		stat_text += str(p) + '% of special shifts scdeduled were the staph\'s  '+ get_counting_string(rank + 1) + ' choice.\n' 
	stat_text += '-------------------------------------------------------------\n'
	special_shift_totals = {}
	for name in sorted(by_names):
		stat_text += name + ' has: '
		total = 0
		for rank in sorted(by_names[name]):
			for shift in by_names[name][rank]:
				stat_text += shift.title + ', '
		stat_text = stat_text[:-2] + '\n'
		for rank in sorted(by_names[name]):
			total += len(by_names[name][rank])
			preference = get_counting_string(rank + 1)
			p = float(total / float(rank + 1)) *100
			stat_text += '	'+ str(len(by_names[name][rank])) + ' of ' + str(preference) + ' prefrence scheduled, ' + str(p)+ '% in top ' + str(rank + 1) + '\n'
		if total in special_shift_totals:
			special_shift_totals[total]  += 1
		else:
			special_shift_totals[total] = 1
		if total == 0:
			warnings += '****** ' + name.upper() + ' HAS NO SPECIAL SHIFTS! ******\n'

	totals = '-------------------------------------------------------------\n'
	for total in special_shift_totals:
		totals += str(special_shift_totals[total]) + ' staphers with ' + str(total) + ' special shifts.\n'
	totals += '-------------------------------------------------------------\n'
	stat_text = warnings + totals + stat_text
	print stat_text
	directory = '../output/2017/masters'
	if not os.path.exists(directory):
		os.makedirs(directory)
	filename = directory + '/special_shift_stats.txt'
	new_file = open(filename, 'w')
	new_file.write(stat_text)
	new_file.close()

def special_shifts(staph, shifts):
	print 'Scheduling Special Shifts...'
	cushion = 2
	special_shifts_to_cover = list(shifts['all-non-flexible-shifts']) #Needs to be updated for flexible shifts
	all_preferences = shifts['non-flexible'] #Needs to be updated for flexible shifts
	max_special_shifts = int(len(special_shifts_to_cover)/ len(staph['all-staph'])) + cushion
	special_shift_types = shifts['special']['non-flexible']
	stapher_queue = sorted(staph['all-staph'], cmp=by_class_year_and_summers)
	name_to_count = {}
	for stapher in stapher_queue:
		name_to_count[stapher.name] = 0
	shift_to_remove = []
	for shift in special_shifts_to_cover:
		if shift.covered:
			name_to_count[shift.stapher.name] += 1
			shift_to_remove.append(shift)
	for shift in shift_to_remove:
		special_shifts_to_cover.remove(shift)
	while len(special_shifts_to_cover) > 0 and len(stapher_queue) > 0:
		stapher = stapher_queue[0]
		ranked_shift = get_top_ranked_special_shift(shifts,stapher)
		if ranked_shift in special_shifts_to_cover and name_to_count[stapher.name] < max_special_shifts:
			stapher.add_shift(ranked_shift)
			stapher_queue.remove(stapher)
			stapher_queue.append(stapher)
			special_shifts_to_cover.remove(ranked_shift)
			name_to_count[stapher.name] += 1
		else:
			stapher_queue.remove(stapher)
	count = 0
	while len(special_shifts_to_cover) > 0:
		for stapher in staph['all-staph']:
			if name_to_count[stapher.name] == count:
				for shift in special_shifts_to_cover:
					preference = get_preference_of_shift(shift,all_preferences,shifts)
					if stapher.free_during_shift(shift) and name_to_count[stapher.name] < max_special_shifts and stapher_has_no_shift_of_type(stapher,preference,shifts):
						stapher.add_shift(shift)
						name_to_count[stapher.name] += 1
						special_shifts_to_cover.remove(shift)
		max_special_shifts = max_special_shifts + count
		count += 1
	# schedule_special_shifts(staph, shifts)
	
def meal_shifts(staph,shifts):
	for meal in shifts['meal']:
		print meal
		meal_shifts_to_fill = shifts[meal]
		while len(meal_shifts_to_fill) != 0:
			staphers = list(staph['all-staph'])
			meal_shifts_to_fill = shifts[meal]
			print len(meal_shifts_to_fill)
			tries_for_stapher = 0
			tried_shifts = []
			while len(meal_shifts_to_fill) > 0 and not len(staphers) == 0:
				# print len(staphers), 'to cover',len(meal_shifts_to_fill),'shifts.'
				stapher = staphers[0]
				meal_shift = meal_shifts_to_fill[random.randint(0,len(meal_shifts_to_fill) - 1)]
				while meal_shift in tried_shifts:
					meal_shift = meal_shifts_to_fill[random.randint(0,len(meal_shifts_to_fill) - 1)]
				tried_shifts.append(meal_shift)
				if stapher.free_during_shift(meal_shift):
					stapher.add_shift(meal_shift)
					staphers.remove(stapher)
					staphers.append(stapher)
					meal_shifts_to_fill.remove(meal_shift)
					tried_shifts = []
				# print '	tried',len(tried_shifts),'out of',len(meal_shifts_to_fill),'shifts'
				if len(tried_shifts) == len(meal_shifts_to_fill):
					staphers.remove(stapher)
					tried_shifts = []
			print '	',len(meal_shifts_to_fill),'uncovered'
			for uncovered_meal_shift in meal_shifts_to_fill:
				print uncovered_meal_shift
				# print get_who_is_free_during_shift(staph,uncovered_meal_shift)
		
def get_readable_time(time):
	hour = int(time)
	minute = str(time - hour)
	if minute == '0.0' or minute == '0':
		minute = ':00'
	elif minute == '0.25':
		minute = ':15'
	elif minute == '0.33':
		minute = ':20'
	elif minute == '0.5':
		minute = ':30'
	elif minute == '0.66':
		minute = ':40'
	elif minute == '0.75':
		minute = ':45'
	if hour < 12 or hour == 24:
		suffix = 'am'
	else:
		suffix = 'pm'
	if hour > 12:
		hour = hour - 12
	if hour == 0:
		hour = 12
	return str(hour) + minute + suffix

def make_schedule_txt_files(staph,directory):
	days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday','Saturday']
	for stapher in staph['all-staph']:
		schedule_txt = ''
		schedule = stapher.schedule.all_shifts
		for day in schedule:
			schedule_txt += str(days[day]) + ','
			for shift in schedule[day]:
				if shift.is_programming:
					programming_tag = 'p'
				else:
					programming_tag = ''
				start = get_readable_time(shift.start)
				end = get_readable_time(shift.end)
				schedule_txt +=  shift.title + '(' + start + '-' + end + ')' + programming_tag + ','
			schedule_txt = schedule_txt[:-1] + '\n'
		schedule_txt = schedule_txt[:-1]
		if not os.path.exists(directory):
			os.makedirs(directory)
		filename = directory + '/' + stapher.name + '.txt'
		new_file = open(filename, 'w')
		new_file.write(schedule_txt)
		new_file.close()

def reset_xl_files(ws):
    merged_cells = ws.merged_cell_ranges
    for day in range(0,7):
		row = 9 + (day * 4)
		time = 6
		while time < 23.5:
			hour = int(time)
			minute = time - hour
			col = 3 + ((hour - 6) * 4) + (minute * 4)
			time_cell = ws.cell(row=row - 2,column=col)
			programming_cell = ws.cell(row=row - 1,column=col)
			cell = ws.cell(row=row,column=col)
			cell.value = ''
			programming_cell.value = ''
			time_cell.value = ''
			cell.fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('ffffff'))
			cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
			cell_coordinance = cell.column + str(cell.row)
			for m_cell in merged_cells:
				m_start = m_cell[:m_cell.index(':')]
				if cell_coordinance == m_start:
					end_coordinance = m_cell[m_cell.index(':') + 1:]
					cell_to_unmerge = cell_coordinance + ':' + end_coordinance
					ws.unmerge_cells(cell_to_unmerge)
			time += 0.25

def get_start_time(shift_txt):
	time_txt = shift_txt[shift_txt.index('('):]
	start_index = time_txt.index('-')
	start_time_txt = time_txt[1:start_index]
	start_time = get_time_from_txt(start_time_txt)
	return start_time

def get_end_time(shift_txt):
	time_txt = shift_txt[shift_txt.index('('):]
	start_index = time_txt.index('-')
	end_time_txt = time_txt[start_index + 1: -1]
	end_time = get_time_from_txt(end_time_txt)
	return end_time

def update_xlsx(src, ws):
	days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday','Saturday']
	schedule_file = open(src)
	csv_schedule_file = csv.reader(schedule_file)
	for line in csv_schedule_file:
		day = days.index(line[0])
		row = 9 + (day * 4)
		for shift_txt in line[1:]:
			is_programming = shift_txt[len(shift_txt) - 1] == 'p'
			if is_programming:
				shift_txt = shift_txt[:-1]
			start_time = get_start_time(shift_txt)
			start_hour = int(start_time)
			start_minute = start_time - start_hour
			end_time = get_end_time(shift_txt)
			end_hour = int(end_time)
			end_minute = end_time - end_hour
			if str(end_minute) == '0.33':
				end_minute = 0.25
			if str(end_minute) == '0.66':
				end_minute = 0.75
			start_col = 3 + ((start_hour - 6) * 4) + (start_minute * 4)
			end_col = 3 + ((end_hour - 6) * 4) + (end_minute * 4) - 1
			cell = ws.cell(row=row,column=start_col)
			shift_title =  shift_txt[:shift_txt.index('(')]
			if ':20' in shift_txt or ':40' in shift_txt:
				cell.value = shift_txt
			else:
				cell.value = shift_title
			# print shift_txt,row,start_col,end_col
			if shift_title == 'Off Day':
				cell.fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('C4C4C4'))
			else:
				curr_col = start_col
				while curr_col <= end_col:
					time_cell = ws.cell(row=row - 2,column=curr_col)
					programming_cell = ws.cell(row=row - 1,column=curr_col)
					if is_programming:
						programming_cell.value = '1'
					else:
						time_cell.value = '1'
					curr_col += 1
					programming_cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
			cell.alignment = Alignment(shrinkToFit=True, wrapText=True, horizontal='center',vertical='center')
			cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
			ws.merge_cells(start_row=row,start_column=start_col,end_row=row,end_column=end_col)
	schedule_file.close()

def get_xl_group(position):
	if position in ['volleyball','yoga','tennis','climbing','sailing']:
		return 'athletics'
	elif position in ['arts','crafts','music','theater']:
		return 'arts'
	elif position in ['hiking','naturalist','kids-naturalist']:
		return 'outdoors'
	elif position in ['office','photo']:
		return 'opho'
	elif position in ['iic','kgc','staph-d']:
		return 'managers'
	else:
		return position

def fill_xl_with_schedules(staph):
	for stapher in staph['all-staph']:
		positions = ''
		for position in stapher.positions:
			positions += position + ' & '
		positions = positions[:-3]
		group = get_xl_group(stapher.positions[0])
		filename = '../output/2017/schedules/xl_files/' + group + '-schedules.xlsx'
		name = stapher.name
		print 'Loading ' + group + ' workbook for ' + name + '...'
		wb = load_workbook(filename)
		if name not in wb.sheetnames:
			template = wb.get_sheet_by_name('TEMPLATE')
			new = wb.copy_worksheet(template)
			new.title = name
			name_cell = new['A2']
			name_cell.value = 'Name: ' + name
			position_cell = new['AE2']
			position_cell.value = 'Position: ' + positions.upper()
			print 'MADE A NEW WORKSHEET IN', filename, 'FOR', name
		else:
			ws = wb.get_sheet_by_name(name)
			position_cell = ws['AE2']
			position_cell.value = 'Position: ' + positions.upper()
			reset_xl_files(ws)
			txt_file = '../output/2017/schedules/txt_files/'+ name + '.txt'
			update_xlsx(txt_file,ws)
		print 'Saving workbook...'
		wb.save(filename)

def duplicate_xl_temp(wb,name):
	print 'MAKEING NEW WORKSHETT CALLED', name
	template = wb.get_sheet_by_name('TEMPLATE')
	new = wb.copy_worksheet(template)
	new.title = name
	title_cell = new['B2']
	title_cell.value = name

def add_key_or_value_to_dict(key, value, dictionary):
	if key in dictionary:
		dictionary[key] += value
	else:
		dictionary[key] = value

def group_masters(shifts,staph):
	masters = ['ski-dock','manager','athletics','arts','nature']
	no_individual_masters = ['icom','boat-dock-coordinator','sdh','ski-dock','iic','kgc','yoga','climbing','volleyball','tennis','sailing','arts','crafts','music','theater','hiking','naturalist','kids-naturalist']
	for position in staph:
		if position in shifts:
			if position not in no_individual_masters:
				shifts[ position + '-positions-master'] = shifts['all-' + position + '-shifts']
				masters.append(position)
			if position in ['sdh','ski-dock']:
				add_key_or_value_to_dict('ski-dock-positions-master', shifts['all-' + position + '-shifts'], shifts)
			if position in ['iic','kgc']:
				add_key_or_value_to_dict('manager-positions-master', shifts['all-' + position + '-shifts'], shifts)
			if position in ['yoga','climbing','volleyball','tennis','sailing']:
				add_key_or_value_to_dict('athletics-positions-master', shifts['all-' + position + '-shifts'], shifts)
			if position in ['arts','crafts','music','theater']:
				add_key_or_value_to_dict('arts-positions-master', shifts['all-' + position + '-shifts'], shifts)
			if position in ['hiking','naturalist','kids-naturalist']:
				add_key_or_value_to_dict('outdoors-positions-master', shifts['all-' + position + '-shifts'], shifts)
			if position in ['boat-dock-coordinator']:
				add_key_or_value_to_dict('dock-dock-master', shifts['all-' + position + '-shifts'], shifts)
	special_shifts = shifts['non-flexible']
	for key in special_shifts:
		shift_type = key[:-7]
		if shift_type in ['boatster','astrocruise']:
			add_key_or_value_to_dict('boatster-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['basketball','rowing','staph-v-guest-volleyball','archery','mountain-biking','exercise','fishing','frisbee-golf']:
			add_key_or_value_to_dict('sporting-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['bread-bake','shabbat','dance','meditation','language-hour']:
			add_key_or_value_to_dict('culture-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['cherry-house','social-hike']:
			add_key_or_value_to_dict('social-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['dinner-hobart','lunch-hobart','breakfast-hobart']:
			add_key_or_value_to_dict('hobart-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['campfire','disco-bingo','improv','music-hour','play','games-day','trivia']:
			add_key_or_value_to_dict('performance-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['faculty-speaker-setup','guest-to-guest','discussion','book-discussion','womens-breakfast']:
			add_key_or_value_to_dict('discussion-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['fountain']:
			add_key_or_value_to_dict('fountain-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['lifeguarding']:
			add_key_or_value_to_dict('lifeguarding-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['greeting','parking']:
			add_key_or_value_to_dict('saturday-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['linen-bar','balderdash']:
			add_key_or_value_to_dict('random-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['omelet-bar','breakfast-hobart']:
			add_key_or_value_to_dict('breakfast-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['lunch-hobart']:
			add_key_or_value_to_dict('lunch-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['wine','dinner-hobart']:
			add_key_or_value_to_dict('dinner-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['workshop','music-special-lessons']:
			add_key_or_value_to_dict('workshop-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['musnoo','hoohune','egg-drop','kids-carnival','kids-t-show','puppet-show','story-hour']:
			add_key_or_value_to_dict('kids-special-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['town-run']:
			add_key_or_value_to_dict('town-run-master', shifts[shift_type + '-shifts'], shifts)
		elif shift_type in ['plants']:
			add_key_or_value_to_dict('maintenance-master', shifts[shift_type + '-shifts'], shifts)
	meal_shifts = shifts['meal']
	for key in meal_shifts:
		meal_type = key[:-7]
		add_key_or_value_to_dict('meal-master', shifts[meal_type], shifts)
		for meal in ['breakfast','lunch','dinner']:
			add_key_or_value_to_dict('meal-master', shifts[ meal + '-hobart-shifts'], shifts)

def get_total_overlapping_shifts(shift,shifts_to_populate):
	total_overlapping_shifts = []
	for other_shift in shifts_to_populate:
		if shift != other_shift:
			if shift.at_same_time(other_shift):
				if shift.title == other_shift.title:
					total_overlapping_shifts.append(other_shift)
	return total_overlapping_shifts

def get_partial_overlapping_shifts(shift,shifts_to_populate):
	total_partial_shifts = []
	for other_shift in shifts_to_populate:
		if shift != other_shift:
			if not shift.at_same_time(other_shift) and shift.time_overlaps_with(other_shift):
				total_partial_shifts.append(other_shift)
	return total_partial_shifts

def get_row_end_from_overlap(row_start,overlap):
	if overlap == 0:
		row_end = row_start + 5
	elif overlap < 3:
		row_end = row_start + (3 - overlap)
	else:
		row_end = row_start
	return row_end

def get_minute(time):
	minute = time - int(time)
	if str(minute) == '0.33':
		minute = 0.25
	if str(minute) == '0.66':
		minute = 0.75
	return minute

def get_first_name(name):
	return name[:name.index(' ')]

def populate_master(wb,name,shifts_to_populate):
	# print name, len(shifts_to_populate)
	ws = wb.get_sheet_by_name(name)
	for shift in shifts_to_populate:
		total_overlapping_shifts = get_total_overlapping_shifts(shift,shifts_to_populate)
		partial_overlapping_shifts = get_partial_overlapping_shifts(shift,shifts_to_populate)
		row_start = 7 + (shift.day * 7)
		row_end = get_row_end_from_overlap(row_start, len(partial_overlapping_shifts))
		
		# total overlapping shifts
		start_hour = int(shift.start)
		start_minute = get_minute(shift.start)
		end_hour = int(shift.end)
		end_minute = get_minute(shift.end)
		start_col = 4 + ((start_hour - 6) * 4) + (start_minute * 4)
		end_col = 4 + ((end_hour - 6) * 4) + (end_minute * 4) - 1
		cell = ws.cell(row=row_start,column=start_col)
		# print shift
		if len(total_overlapping_shifts) > 0:
			title = get_first_name(shift.stapher.name) + ','
			for total_overlap_shift in total_overlapping_shifts:
				title += get_first_name(total_overlap_shift.stapher.name) + ', '
				shifts_to_populate.remove(total_overlap_shift)
			title = title[:-2]
		else:
			title = get_first_name(shift.stapher.name)
		title = shift.title + ': ' + title
		cell.value = title
		cell.alignment = Alignment(shrinkToFit=True, wrapText=True, horizontal='center',vertical='center')
		cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		# print '		Cell:', shift,row_start,row_end,start_col,end_col,title
		ws.merge_cells(start_row=row_start,start_column=start_col,end_row=row_end,end_column=end_col)
		
		# Partial Overlapping Shift
		for partial_overlapping_shift in partial_overlapping_shifts:
			other_total_lapping_shifts = get_total_overlapping_shifts(partial_overlapping_shift,partial_overlapping_shifts)
			if len(other_total_lapping_shifts) > 0:
				title = get_first_name(partial_overlapping_shift.stapher.name)
				for shift in other_total_lapping_shifts:
					title = get_first_name(shift.stapher.name) + ', '
					other_total_lapping_shifts.remove(shift)
				title = title[:-2]
			else:
				title = get_first_name(partial_overlapping_shift.stapher.name)
			title = partial_overlapping_shift.title + ': ' + title
			row_start = row_end + 1
			row_end = get_row_end_from_overlap(row_start, len(partial_overlapping_shifts))
			start_hour = int(partial_overlapping_shift.start)
			start_minute = get_minute(partial_overlapping_shift.start)
			end_hour = int(partial_overlapping_shift.end)
			end_minute = get_minute(partial_overlapping_shift.end)
			start_col = 4 + ((start_hour - 6) * 4) + (start_minute * 4)
			end_col = 4 + ((end_hour - 6) * 4) + (end_minute * 4) - 1
			overlap_cell = ws.cell(row=row_start,column=start_col)
			overlap_cell.value = title
			overlap_cell.alignment = Alignment(shrinkToFit=True, wrapText=True, horizontal='center',vertical='center')
			overlap_cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
			ws.merge_cells(start_row=row_start,start_column=start_col,end_row=row_end,end_column=end_col)
			shifts_to_populate.remove(partial_overlapping_shift)

def make_masters_from_key(shifts,master_key):
	file = '../output/2017/masters/' + master_key + '.xlsx'
	wb = load_workbook(file)
	for key in shifts:
		if master_key in key:
			name = key[:-(len(master_key) + 1)].upper() + ' MASTER'
			shifts_to_populate = list(shifts[key])
			if name not in wb.sheetnames:
				print 'Updating',file
				duplicate_xl_temp(wb,name)
				wb.save(file)
			else:
				print 'Populating', name
				populate_master(wb,name,shifts_to_populate)
				wb.save(file)

def make_masters(staph,shifts):
	group_masters(shifts,staph)
	make_off_day_xl(staph)
	make_masters_from_key(shifts,'positions-master')
	make_masters_from_key(shifts,'special-master')

def get_who_is_free_during_time(staph,day,start_time,end_time):
	free_staphers = []
	for stapher in staph['all-staph']:
		if stapher.free_during_time(day,start_time,end_time):
			free_staphers.append(stapher)
	return free_staphers		

def get_who_is_free_during_shift(staph,shift):
	return get_who_is_free_during_time(staph, shift.day, shift.start, shift.end)

def get_stapher_by_name(name,all_staph):
	for stapher in all_staph:
		if name in stapher.name:
			return stapher
	print 'No stapher with',name,'in their name...'

def get_free_shift_by_title(title,all_shifts):
	for shift in all_shifts:
		if title == shift.title and not shift.covered:
			return shift
	print 'No shift with',title,'in the title...'

def fill_pre_determined_shifts(staph,shifts):
	austin = get_stapher_by_name('Austin',staph['all-staph'])
	beffa = get_free_shift_by_title('ICOM BEFFA',shifts['all-shifts'])
	austin.add_shift(beffa)
	alida = get_stapher_by_name('Alida',staph['all-staph'])
	chicken = get_free_shift_by_title('Chicken!',shifts['all-shifts'])
	alida.add_shift(chicken)
	thom = get_stapher_by_name('Thomas R',staph['all-staph'])
	chicken = get_free_shift_by_title('Chicken!',shifts['all-shifts'])
	thom.add_shift(chicken)
	jonny = get_stapher_by_name('Jonny',staph['all-staph'])
	clinic = get_free_shift_by_title('Ski Clinic',shifts['all-shifts'])
	jonny.add_shift(clinic)
	gavbot = get_stapher_by_name('Gavin',staph['all-staph'])
	clinic = get_free_shift_by_title('Ski Clinic',shifts['all-shifts'])
	gavbot.add_shift(clinic)
	clinic = get_free_shift_by_title('Wakeboard Clinic',shifts['all-shifts'])
	austin.add_shift(clinic)
	chankins = get_stapher_by_name('Cody',staph['all-staph'])
	clinic = get_free_shift_by_title('Wakeboard Clinic',shifts['all-shifts'])
	chankins.add_shift(clinic)
	halle = get_stapher_by_name('Halle',staph['all-staph'])
	clinic = get_free_shift_by_title('Teen Ski',shifts['all-shifts'])
	halle.add_shift(clinic)
	dylan = get_stapher_by_name('Dylan',staph['all-staph'])
	clinic = get_free_shift_by_title('Teen Ski',shifts['all-shifts'])
	dylan.add_shift(clinic)

	cast = ['Thomas C','Daniel H','Gracie','Leo C','Martin A','McKenzie','Susi']
	while len(cast) > 0:
		stapher = get_stapher_by_name(cast[0],staph['all-staph'])
		play_shift = get_free_shift_by_title('Brothers Grimm',shifts['all-shifts'])
		stapher.add_shift(play_shift)
		cast.remove(cast[0])

def remove_all_shifts(staph):
	for stapher in staph['all-staph']:
		for shift in stapher.all_shifts():
			stapher.remove_shift(shift)

def make_schedules(staph,shifts,erase_schedules,reset_schedules,upload_schedules,shifts_to_generate,save_schedules,make_statistic_files,output_xl_schedules):
	if erase_schedules:
		print '	SCHEDULES ERASED'
		make_schedule_txt_files(staph,'../input/2017/schedules/txt_files')
	elif reset_schedules:
		print '	SCHEDULES RESET'
		fill_pre_determined_shifts(staph,shifts)
		upload_off_days(staph,shifts)
	elif upload_schedules:
		print '	SCHEDULES UPLOADED'
		get_schedules_from_txt_files(staph,shifts)
	if shifts_to_generate[0]:
		print '	GENERATING SCHEDULES W/ OFF DAYS'
		generate_off_days(staph,shifts)
	if shifts_to_generate[1]:
		print '	GENERATING SCHEDULES W/ PROGRAMMING'
		programming(staph,shifts)
	if shifts_to_generate[2]:
		print '	GENERATING SCHEDULES W/ SPECIAL SHIFTS'
		special_shifts(staph, shifts)
	if shifts_to_generate[3]:
		print '	GENERATING SCHEDULES W/ MEAL SHIFTS'
		print 'TODO: need to implement meal shifts'
	if shifts_to_generate[4]:
		# print '	GENERATING SCHEDULES W/ LIFEGUADING SHIFTS'
		print '		need to implement lifeguarding shift automation'
	if shifts_to_generate[5]:
		# print '	GENERATING SCHEDULES W/ BOAT DOCK SHIFTS'
		print '		need to implement boat dock shift automation'
	if shifts_to_generate[6]:
		# print '	GENERATING SCHEDULES W/ FOUNTAIN SHIFTS'
		print '		need to implement fountain shift automation'
	if save_schedules:
		print '	SCHEDULES SAVED'
		make_schedule_txt_files(staph,'../input/2017/schedules/txt_files')
	if make_statistic_files:
		print '	MAKING STATISTICS FILES'
		make_off_day_statistics(staph,shifts)
		make_programming_stats(staph)
		# make_special_shift_statistics(staph,shifts)
	if output_xl_schedules:
		print '	OUTPUTTING XL + STATISTICS'
		make_schedule_txt_files(staph,'../output/2017/schedules/txt_files')
		fill_xl_with_schedules(staph)
		make_masters(staph,shifts)
	# Don't need this, but it is nice to see where I am at the end of each running 
	make_schedule_txt_files(staph,'../output/2017/schedules/txt_files')

if __name__ == "__main__":
	staph_file = '../input/2017/staph.csv'
	shift_dir = '../input/2017/shifts'
	staph = get_staph_from_csv_file(staph_file)
	shifts = get_shifts(shift_dir)
	erase_schedules = False
	reset_schedules = False
	upload_schedules = True
	shifts_to_generate = [False,False,False,False,False,False,False]
	save_schedules = True
	make_statistic_files = True
	output_xl_schedules = False
	make_schedules(staph,shifts,erase_schedules,reset_schedules,upload_schedules,shifts_to_generate,save_schedules,make_statistic_files,output_xl_schedules)








